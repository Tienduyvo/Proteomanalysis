import openpyxl
import string


#################### This script subtract unwated column from the negativ and LB sample ################

################################### read the Triplicate proteingroup table #############################

input_triplicate_workbook = openpyxl.load_workbook("Xsze_proteingroups_Triplicate.xlsx")

input_triplicate_worksheet = input_triplicate_workbook.active

################################### read the LB proteingroup table #####################################

input_lb_workbook = openpyxl.load_workbook("Xsze_proteingroups_in_LB.xlsx")

input_lb_worksheet = input_lb_workbook.active

#################################### read the proteingroup table from the negative sample ##############

input_negative_workbook = openpyxl.load_workbook("Xsze_proteingroups_negative.xlsx")

input_negative_worksheet = input_negative_workbook.active

##################################### read the outputtable #############################################

output_workbook = openpyxl.load_workbook('Xsze_proteingroups_Triplicate_reduced.xlsx')

output_worksheet = output_workbook.active


##################################### find out the numbers of rows ######################################
lb_length = input_lb_worksheet.max_row

neg_length = input_negative_worksheet.max_row

triplicate_length = input_triplicate_worksheet.max_row

i = 3
j = 3
k = 3

alphabet = list(string.ascii_uppercase)

while i <= triplicate_length:
	
	j = 3
	while j <= lb_length:

		k = 3

		while k <= neg_length:

			if input_triplicate_worksheet['O'+ str(i)].value == input_lb_worksheet['O'+str(j)].value or input_triplicate_worksheet['O'+ str(i)].value == input_negative_worksheet['O'+str(k)].value:
				output_worksheet['O'+ str(i)].value = ''
				print(str(i)+' '+str(j)+' '+str(k))
				break

			k = k + 1

		j = j + 1

	i = i + 1 

	
		

output_workbook.save('Xsze_proteingroups_Triplicate_reduced.xlsx')	 